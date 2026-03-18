import pandas as pd 
from openpyxl import load_workbook

# Carregar o workbook
wb = load_workbook("pedidos_griffes_ficticias.xlsx")

# Trocar nome da planilha
ws = wb['Pedidos']
ws.title = "PANELA"
wb.save("pedidos_griffes_ficticias.xlsx")

    
# importar o arquivo Excel
df = pd.read_excel("pedidos_griffes_ficticias.xlsx")

# Excluindo colunas
colunas = ['Data Entrega Prevista', 'Data Entrega Real', 'Documento Cliente', 'Transportadora', 'Condição de Pagamento']
df = df.drop(columns=colunas)

# Alterando o nome das planilhas (usa openpyxl, pois xlsxwriter não aceita modo append)
with pd.ExcelWriter("pedidos_griffes_ficticias.xlsx", engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='PANELA', index=False)
    df.to_excel(writer, sheet_name='CASTRO', index=False)
    
df_novo = pd.read_excel("pedidos_griffes_ficticias.xlsx", sheet_name=['PANELA', 'CASTRO'])

df_panela = df_novo['PANELA']
df_castro = df_novo['CASTRO']

# Filtrando Griffes Panela
griffes_panela = ['Aura & Co Fem', 'Aura & Co Masc', 'Vanguardia', 'L\'Éclat Fem', 'L\'Éclat Masc']
linhas_excluir_panela = ['Malha', 'Malha Black', 'Moletom']
df_panela = df_panela.drop(df_panela[df_panela['Linha'].isin(linhas_excluir_panela)].index)
# Mantém apenas Tricot Feminino e Underwear Masculino na aba PANELA
filtro_panela = (
    (df_panela['Linha'] != 'Tricot') & (df_panela['Griffe'].isin(['Aura & Co Masc', 'L\'Éclat Masc', 'Vanguardia']))
) | (
    (df_panela['Linha'] != 'Underwear') & (df_panela['Griffe'].isin(['Aura & Co Fem', 'L\'Éclat Fem']))
)

df_panela = df_panela[df_panela['Griffe'].isin(griffes_panela) & filtro_panela]

# Filtrando Griffes Castro
griffes_castro = ['Aura & Co Fem', 'L\'Éclat Fem']
linhaas_castro = ['Malha', 'Malha Black', 'Moletom', 'Underwear']
df_castro = df_castro[df_castro['Griffe'].isin(griffes_castro) & df_castro['Linha'].isin(linhaas_castro)]

with pd.ExcelWriter("pedidos_griffes_ficticias.xlsx", engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df_panela.to_excel(writer, sheet_name='PANELA', index=False)
    df_castro.to_excel(writer, sheet_name='CASTRO', index=False)


from openpyxl.utils import get_column_letter


wb = load_workbook('pedidos_griffes_ficticias.xlsx')
ws = wb.active

# Filtros automáticos
for ws in wb.worksheets:
    if ws.max_row > 1 and ws.max_column > 0:
        first_col = 1
        last_col = ws.max_column
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}1"



# Ajustar a largura das colunas para todas as sheets
for ws in wb.worksheets:
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 2



# Adicionando bordas
from openpyxl.styles import Border, Side

borda = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)


for ws in wb.worksheets:
    for linha in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for celula in linha:
            celula.border = borda



# Adicionando cor de fundo no cabeçalho somente
from openpyxl.styles import PatternFill

fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')

for ws in wb.worksheets:
    # Preenche apenas a primeira linha (cabeçalho), até a última coluna com dados
    header_row = 1
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value is not None:
            cell.fill = fill

# Adicionando um (ou mais) 0 à esquerda

for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Supondo que o ID do pedido esteja na coluna A
        for cell in row:
            if cell.value is not None:
                cell.number_format = '000000'  # Formato para exibir 6 dígitos com zeros à esquerda
wb.save('Carteira_Fictícia.xlsx')

