import pandas as pd
from openpyxl import *
from openpyxl.utils import get_column_letter


wb = load_workbook('pedidos_griffes_ficticias.xlsx')
ws = wb.active

# Filtros automáticos
for ws in wb.worksheets:
    ws.auto_filter.ref = 'A1:J1'



# Ajustar a largura das colunas para todas as sheets
for ws in wb.worksheets:
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[get_column_letter(col)].width = value + 2



# Adicionando bordas
from openpyxl.styles import Border, Side

borda = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)

max_linhas = ws.max_row
max_colunas = ws.max_column

for ws in wb.worksheets:
    for linha in ws.iter_rows(min_row=1, max_row=max_linhas, min_col=1, max_col=max_colunas):
        for celula in linha:
            celula.border = borda



# Adicionando cor de fundo
from openpyxl.styles import PatternFill

fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')

for ws in wb.worksheets:
    for row in ws["A1:AR1"]:
        for cell in row:
            cell.fill = fill

wb.save('Carteira_Fictícia.xlsx')